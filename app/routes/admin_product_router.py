import io
import re
import unicodedata
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, File, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook
from pydantic import BaseModel

from app.config import IMAGE_EMBEDDING_MODEL, IMAGE_EMBEDDING_PRETRAINED
from app.database.connection import database_connection
from app.services.product_sync_service import product_sync_manager


router = APIRouter(prefix="/admin/products", tags=["Product Admin"])
PAGE_PATH = Path(__file__).resolve().parent.parent / "static" / "product_admin.html"
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
SKU_HEADERS = {"sku", "ma san pham", "product code", "product sku", "variant sku"}


class SkuImportRequest(BaseModel):
    skus: list[str]


def _normalize_header(value: object) -> str:
    normalized = unicodedata.normalize("NFD", str(value or "").strip().casefold())
    normalized = "".join(
        character for character in normalized
        if unicodedata.category(character) != "Mn"
    ).replace("đ", "d")
    return re.sub(r"[\s_-]+", " ", normalized).strip()


def _read_excel_skus(content: bytes) -> list[str]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("File Excel không hợp lệ hoặc bị hỏng.") from error
    rows = workbook.active.iter_rows(values_only=True)
    header = next((row for row in rows if any(str(v or "").strip() for v in row)), None)
    if not header:
        raise ValueError("File Excel không có dữ liệu.")
    headers = [_normalize_header(value) for value in header]
    sku_index = next((i for i, name in enumerate(headers) if name in SKU_HEADERS), None)
    if sku_index is None:
        raise ValueError("Excel phải có cột SKU, Mã sản phẩm hoặc Product Code.")
    skus = [
        str(row[sku_index]).strip()
        for row in rows
        if sku_index < len(row) and row[sku_index] is not None and str(row[sku_index]).strip()
    ]
    if not skus:
        raise ValueError("Cột SKU không có mã sản phẩm.")
    return skus


def _start_job(skus: list[str], tasks: BackgroundTasks) -> dict:
    try:
        job = product_sync_manager.create(skus)
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error
    tasks.add_task(product_sync_manager.run, job.id)
    return job.public()


@router.get("", response_class=HTMLResponse)
def page():
    return HTMLResponse(PAGE_PATH.read_text(encoding="utf-8"), headers={"Cache-Control": "no-store"})


@router.post("/api/import-skus")
def import_skus(data: SkuImportRequest, background_tasks: BackgroundTasks):
    return _start_job(data.skus, background_tasks)


@router.post("/api/import-excel")
async def import_excel(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if Path(file.filename or "").suffix.casefold() != ".xlsx":
        raise HTTPException(status_code=415, detail="Chỉ hỗ trợ file .xlsx.")
    content = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File vượt quá 5 MB.")
    try:
        return _start_job(_read_excel_skus(content), background_tasks)
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error


@router.get("/api/jobs/{job_id}")
def job(job_id: str):
    result = product_sync_manager.get(job_id)
    if not result:
        raise HTTPException(status_code=404, detail="Không tìm thấy tác vụ.")
    return result


@router.get("/api/catalog")
def catalog(
    search: str = Query("", max_length=100),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
):
    keyword = search.strip()
    where = ""
    search_params: list[object] = []
    if keyword:
        where = "WHERE p.product_code ILIKE %s OR p.title ILIKE %s OR p.product_type ILIKE %s"
        pattern = f"%{keyword}%"
        search_params = [pattern, pattern, pattern]
    with database_connection() as connection:
        total = connection.execute(
            f"SELECT COUNT(*) count FROM products p {where}", search_params
        ).fetchone()["count"]
        rows = connection.execute(
            f"""
            SELECT p.product_code, p.title, p.product_type, p.status, p.updated_at,
                   COUNT(DISTINCT pv.id) variant_count,
                   COUNT(DISTINCT pi.id) FILTER (WHERE pi.is_active) image_count,
                   COUNT(DISTINCT pi.id) FILTER (
                       WHERE pi.is_active AND COALESCE(pi.local_path, '') <> ''
                   ) local_image_count,
                   COUNT(DISTINCT pie.product_image_id) embedding_count,
                   STRING_AGG(DISTINCT pv.color, ', ' ORDER BY pv.color) colors
            FROM products p
            LEFT JOIN product_variants pv ON pv.product_id = p.id
            LEFT JOIN product_images pi ON pi.product_id = p.id
            LEFT JOIN product_image_embeddings pie
              ON pie.product_image_id = pi.id
             AND pie.model_name = %s AND pie.pretrained_name = %s
            {where}
            GROUP BY p.id ORDER BY p.updated_at DESC, p.product_code
            LIMIT %s OFFSET %s
            """,
            [IMAGE_EMBEDDING_MODEL, IMAGE_EMBEDDING_PRETRAINED, *search_params, limit, offset],
        ).fetchall()
    products = []
    for row in rows:
        item = dict(row)
        item["updated_at"] = item["updated_at"].isoformat() if item["updated_at"] else None
        item["ai_ready"] = (
            item["status"] == "ACTIVE"
            and int(item["local_image_count"] or 0) > 0
            and int(item["embedding_count"] or 0) >= int(item["local_image_count"] or 0)
        )
        products.append(item)
    return {
        "total": int(total), "limit": limit, "offset": offset,
        "embedding_model": IMAGE_EMBEDDING_MODEL, "products": products,
    }
